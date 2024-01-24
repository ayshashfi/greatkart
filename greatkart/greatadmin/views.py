from django.contrib import messages,auth
from django.shortcuts import render,redirect,HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from accounts.models import *
from greatadmin.models import *
from accounts.models import Account
from django.core.paginator import Paginator
from datetime import datetime,date


from orders.models import Notification
from django.db.models import Sum
from orders.models import Order, OrderProduct



def admin_login(request):
    if request.method=='POST':
        email=request.POST['email']
        password=request.POST['password']
        
        user=auth.authenticate(email=email,password=password)
        
        if email.strip()=='' or password.strip()=='':
            messages.error(request,'fields cannot be empty!')
            return redirect('admin_login')
        
        if user is not None:
            if user.is_active:
               if user.is_superadmin:
                    auth.login(request,user)
                    return redirect('admin_dashboard')
               else:
                    messages.warning(request,'Sorry only admin is allowed to login! ')
                    return redirect('admin_login')
            else:
               messages.warning(request,"Your account has been blocked!")
               return redirect('admin_login')

        else:
            messages.error(request,'Invalid username or password!')
            return redirect('admin_login')
    return render(request,'admin/admin_login.html')

   



@login_required(login_url='admin_login')
def admin_dashboard(request):
    if not request.user.is_superadmin:
        return redirect('admin_login')
    notifications = Notification.objects.all().order_by('-timestamp')[:5]
    today_notifications = [notification for notification in notifications if notification.timestamp.date() == date.today()]

    # Sales data for the chart
    sales_data = OrderProduct.objects.values('order__created_at__date').annotate(total_sales=Sum('product_price')).order_by('-order__created_at__date')
    categories = [item['order__created_at__date'].strftime('%d/%m') for item in sales_data]
    sales_values = [item['total_sales'] for item in sales_data]

    # Return data for the chart
    return_data = OrderProduct.objects.filter(order__status__in=["Returned", "Cancelled"]).values('order__created_at__date').annotate(total_returns=Sum('product_price')).order_by('-order__created_at__date')
    return_values = [item['total_returns'] for item in return_data]

    # Recent orders
    orders = Order.objects.order_by('-created_at')[:10]

    # Total sales and earnings
    try:
        total_sales = Order.objects.aggregate(Sum('order_total'))['order_total__sum'] or 0
        total_sales_rounded = round(total_sales, 2)
    except:
        total_sales = 0

    try:
        total_earnings = Order.objects.filter(status='Completed').aggregate(Sum('order_total'))['order_total__sum'] or 0
    except:
        total_earnings = 0

    # Status counts
    status_pending_count = Order.objects.filter(status='New').count()
    status_delivery_count = Order.objects.filter(status='Completed').count()
    status_cancel_count = Order.objects.filter(status='Cancelled').count()
    status_return_count = Order.objects.filter(status='Returned').count()
    total_status_count = status_delivery_count + status_cancel_count + status_return_count

    # Calculate percentages
    status_delivery = (status_delivery_count / total_status_count) * 100 if total_status_count != 0 else 0
    status_cancel = (status_cancel_count / total_status_count) * 100 if total_status_count != 0 else 0
    status_return = (status_return_count / total_status_count) * 100 if total_status_count != 0 else 0

    context = {
        'delivery_count': status_delivery_count,
        'cancel_count': status_cancel_count,
        'pending_count': status_pending_count,
        'totalsale': total_sales,
        'total_sales_rounded':total_sales_rounded,
        'totalearnings': total_earnings,
        'status_delivery': status_delivery,
        'status_cancel': status_cancel,
        'status_return': status_return,
        'orders': orders,
        'categories': categories,
        'sales_values': sales_values,
        'return_values': return_values,
        'notifications': notifications,
        'today_notifications':today_notifications,
    }

    return render(request, 'admin/admin_dashboard.html', context)

    



@login_required(login_url='admin_login')
def usermanagement(request):
    if not request.user.is_superadmin:
        return redirect('admin_login')
    users = Account.objects.filter(is_superadmin=False).order_by('id')
    p=Paginator(users,6)
    page=request.GET.get('page')
    user_page=p.get_page(page)
    page_nums='a'*user_page.paginator.num_pages
    notifications = Notification.objects.all().order_by('-timestamp')[:5]
    today_notifications = [notification for notification in notifications if notification.timestamp.date() == date.today()]
    context={
        'users':users,
        'user_page':user_page,
        'page_nums':page_nums,
        'notifications':notifications,
        'today_notifications':today_notifications,
        
        }
    return render(request,'admin/usermanagement.html',context)



@login_required(login_url='admin_login')   
def blockuser(request,user_id):
    if not request.user.is_superadmin:
        return redirect('admin_login')
    user =Account.objects.get(id=user_id)
    if user.is_active:
        user.is_active = False
        user.save()  
    else:
        user.is_active = True
        user.save()
    return redirect('usermanagement')



@login_required(login_url='admin_login')
def sales_report(request):
    if not request.user.is_superadmin:
        return redirect('admin_login')

    if request.method == 'GET':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if start_date and end_date:
            # Filter orders based on the selected date range
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if start_date > end_date:
                messages.error(request, "Start date must be before end date.")
                return redirect('sales_report')
            if end_date > date.today():
                messages.error(request, "End date cannot be in the future.")
                return redirect('sales_report')

            orders = Order.objects.filter(created_at__date__range=(start_date, end_date))
            recent_orders = orders.order_by('-created_at')
        else:
            # If no date range is selected, fetch recent 10 orders
            recent_orders = Order.objects.order_by('-created_at')[:10]
            orders = Order.objects.all()

        # Calculate total sales and total orders
        total_sales = round(sum(order.order_total for order in orders), 2)
        total_orders = orders.count()
        
        notifications = Notification.objects.all().order_by('-timestamp')[:5]
        today_notifications = [notification for notification in notifications if notification.timestamp.date() == date.today()]

        # Calculate sales by status
        sales_by_status = {
            'New': orders.filter(status='New').count(),
            'Accepted': orders.filter(status='Accepted').count(),
            'Completed': orders.filter(status='Completed').count(),
            'Cancelled': orders.filter(status='Cancelled').count(),
            'Returned': orders.filter(status='Returned').count(),
        }

        

        # Prepare data for rendering the template
        sales_report = {
            'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
            'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
            'total_sales': total_sales,
            'total_orders': total_orders,
            'sales_by_status': sales_by_status,
            'recent_orders': recent_orders,
            'orders': orders,
           
            
        }

        return render(request, 'admin/sales_report.html', {'sales_report': sales_report, 'notifications':notifications,
            'today_notifications':today_notifications,})



from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Prefetch, Sum
from django.shortcuts import redirect
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from accounts.models import Account
from orders.models import Order, OrderProduct
styles = getSampleStyleSheet()
from reportlab.lib import colors


def generate_sales_table(data):
    # Create a table with data
    table = Table(data)

    # Add style to the table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    return table



def generate_sales_report(orders):
    elements = []

    # Header Information
    elements.append(Paragraph('Sales Report', style=styles['Title']))
    elements.append(Paragraph(str(datetime.now()), style=styles['Normal']))
    elements.append(Paragraph('<br/><br/>', style=styles['Normal']))

    data = [['User', 'Total Price', 'Order Number','Ordered At', 'Product Name', 'Product Price', 'Product Quantity']]
    total_sales = 0
    total_orders = 0

    for order in orders:
        order_items = order.orderproduct_set.all()
        total_sales += order.order_total
        total_orders += 1

        for order_item in order_items:
            data.append([
                order.user.first_name,
                'Rs.' + str(order.order_total),
                order.order_number,
                str(order.created_at.date()),
                order_item.product.product_name,
                'Rs.' + str(order_item.product_price),
                order_item.quantity,
            ])

    # Generate and add the table to elements
    table = generate_sales_table(data)
    elements.append(table)

    elements.append(Paragraph('<br/><br/>', style=styles['Normal']))
    # Add Total Sales and Total Orders
    elements.append(Paragraph(f'Total Sales: Rs. {total_sales}', style=styles['Normal']))
    elements.append(Paragraph(f'Total Orders: {total_orders}nos.', style=styles['Normal']))

    return elements


def generate_pdf(request):
    if not request.user.is_superadmin:
        return redirect('admin_login')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=SalesReport' + str(datetime.now()) + '.pdf'
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    orders = Order.objects.all().prefetch_related(
    Prefetch('orderproduct_set', 
             queryset=OrderProduct.objects.select_related('product'),
             to_attr='order_items')
)

    elements = generate_sales_report(orders)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response



import openpyxl
from openpyxl.styles import Font, PatternFill

def generate_sales_report_excel(orders):
    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header Information
    ws['A1'] = 'Sales Report'
    ws['A2'] = str(datetime.now())

    # Set styles for header
    header_style = Font(size=14, bold=True)
    ws['A1'].font = header_style
    ws['A2'].font = header_style

    # Write header row
    header_row = ['User', 'Total Price', 'Order Number', 'Ordered At', 'Product Name', 'Product Price', 'Product Quantity']
    ws.append(header_row)

    total_sales = 0
    total_orders = 0

    for order in orders:
        order_items = order.orderproduct_set.all()
        total_sales += order.order_total
        total_orders += 1

        for order_item in order_items:
            data_row = [
                order.user.first_name,
                'Rs.' + str(order.order_total),
                order.order_number,
                str(order.created_at.date()),
                order_item.product.product_name,
                'Rs.' + str(order_item.product_price),
                order_item.quantity,
            ]
            ws.append(data_row)

    # Apply styles to the header row
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
        cell.font = Font(color="FFFFFF")

    # Add Total Sales and Total Orders
    total_sales_cell = f'Total Sales: Rs. {total_sales}'
    total_orders_cell = f'Total Orders: {total_orders}nos.'

    ws.append([])
    ws.append([total_sales_cell])
    ws.append([total_orders_cell])

    return wb


from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
import io


def generate_excel(request):
    # Ensure the user is authorized (you may have your own authorization logic)
    if not request.user.is_superadmin:
        return redirect('admin_login')

    # Your existing code to fetch and process data
    orders = Order.objects.all().prefetch_related(
        Prefetch('orderproduct_set', 
                queryset=OrderProduct.objects.select_related('product'),
                to_attr='order_items')
    )

    workbook = Workbook()

    # Create a new worksheet
    sheet = workbook.active

    # Add headers to the worksheet
    headers = ['User', 'Order Number', 'Ordered At', 'Product Name', 'Total Price', 'Product Quantity']
    sheet.append(headers)

    # Add data to the worksheet
    total_sales = 0
    total_orders = 0

    for order in orders:
        order_items = order.order_items
        for order_item in order_items:
            sheet.append([
                order.user.first_name,
                order.order_number,
                str(order.created_at.date()),
                order_item.product.product_name,
                'Rs.' + str(order.order_total),
                order_item.quantity,
            ])

            total_sales += order.order_total
            total_orders += 1

    # Add Total Sales and Total Orders
    sheet.append([])  # Add an empty row for spacing
    sheet.append([f'Total Sales: Rs. {total_sales}'])
    sheet.append([f'Total Orders: {total_orders} nos.'])

    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=SalesReport_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'

    # Save the workbook to BytesIO buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Write the BytesIO buffer to the response
    response.write(buffer.read())
    buffer.close()

    return response


from django.http import JsonResponse
from django.views import View
from django.db.models import Count, Sum

class GetChartDataView(View):
    def get(self, request, *args, **kwargs):
        option = request.GET.get('option', 'monthly')

        # Fetch data based on the selected option (monthly/yearly)
        if option == 'monthly':
            # Your logic to fetch monthly data
            monthly_data = self.fetch_monthly_data()
            data = {
                'labels': monthly_data['labels'],
                'datasets': [
                    {'label': 'Total Orders', 'data': monthly_data['order_count']},
                    {'label': 'Total Sales', 'data': monthly_data['total_sales']}
                ]
            }
        elif option == 'yearly':
            # Your logic to fetch yearly data
            yearly_data = self.fetch_yearly_data()
            data = {
                'labels': yearly_data['labels'],
                'datasets': [
                    {'label': 'Total Orders', 'data': yearly_data['order_count']},
                    {'label': 'Total Sales', 'data': yearly_data['total_sales']}
                ]
            }
        else:
            data = {}

        return JsonResponse(data)

    def fetch_monthly_data(self):
        # Example logic to fetch monthly data from the Order model
        labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul']

        monthly_data = Order.objects.filter(created_at__month__in=range(1, 8)).values('created_at__month').annotate(
            order_count=Count('id'),
            total_sales=Sum('order_total')
        )

        data = {
            'labels': labels,
            'order_count': [month['order_count'] for month in monthly_data],
            'total_sales': [month['total_sales'] or 0 for month in monthly_data],
        }

        return data

    def fetch_yearly_data(self):
    # Example logic to fetch yearly data from the Order model
        labels = ['2024', '2023', '2022', '2021']

        yearly_data = Order.objects.filter(created_at__year__in=labels).values('created_at__year').annotate(
        order_count=Count('id'),
        total_sales=Sum('order_total')
    )

        data = {
        'labels': labels,
        'order_count': [year['order_count'] for year in yearly_data],
        'total_sales': [year['total_sales'] or 0 for year in yearly_data],
    }

        return data



@login_required(login_url='admin_login')
def admin_logout1(request):
    logout(request)
    return redirect('admin_login')



